from flask import Blueprint, render_template, flash, redirect, url_for, request, jsonify
from flask_login import login_required, current_user
from datetime import datetime
from .models import Transaction, TransactionItem, Sheet, ProductSheet, SheetProduct
import re
from . import db
from .models import Product
from .models import Personnel
from .models import User

views = Blueprint('views', __name__)


def get_sheet_product(sheet_id, product_id):
    from .models import SheetProduct
    return SheetProduct.query.filter_by(sheet_id=sheet_id, product_id=product_id).first()


def is_phone_duplicate(phone, exclude_id=None):
    q = Personnel.query.filter(Personnel.phone == phone)
    if exclude_id:
        q = q.filter(Personnel.id != exclude_id)
    return q.first() is not None


def update_sheet_product_status(sp):
    if sp.stock <= 0:
        sp.stock = 0
        sp.status = "Out of Stock"
    elif sp.stock <= 10:
        sp.status = "Critical"
    else:
        sp.status = "On Stock"


@views.route('/get-active-product-sheet-for-transaction/<int:sheet_id>', methods=['GET'])
@login_required
def get_active_product_sheet_for_transaction(sheet_id):
    sheet = Sheet.query.get(sheet_id)
    if not sheet:
        return jsonify({"product_sheet_id": None})
    product_sheet = ProductSheet.query.filter_by(date=sheet.date).first()
    return jsonify({"product_sheet_id": product_sheet.id if product_sheet else None})


@views.route('/get-all-products', methods=['GET'])
@login_required
def get_all_products():
    products = Product.query.all()
    return jsonify([{'id': p.id, 'name': p.name, 'category': p.category, 'price': p.price} for p in products])


@views.route('/get-personnel-list', methods=['GET'])
@login_required
def get_personnel_list():
    personnel = Personnel.query.order_by(Personnel.name.asc()).all()
    return jsonify({"personnel": [{"id": p.id, "name": p.name, "phone": p.phone, "role": p.role, "status": p.status} for p in personnel]})


@views.route('/')
@login_required
def home():
    today_str = datetime.now().strftime('%m/%d/%Y')
    product_sheets = ProductSheet.query.filter_by(is_active=True).order_by(ProductSheet.date_created.asc()).all()
    if not product_sheets:
        first_psheet = ProductSheet(date=today_str, label=today_str)
        db.session.add(first_psheet)
        db.session.commit()
        product_sheets = [first_psheet]
    active_product_sheet_id = request.args.get('product_sheet_id', product_sheets[0].id, type=int)
    product_sheet = ProductSheet.query.get(active_product_sheet_id)
    if product_sheet:
        sheet_products = SheetProduct.query.filter_by(sheet_id=active_product_sheet_id).join(Product).order_by(Product.name.asc()).all()
    else:
        sheet_products = []
    inventory_items = []
    for sp in sheet_products:
        p = sp.product
        p.stock = sp.stock
        p.status = sp.status
        p.last_modified = sp.last_modified
        inventory_items.append(p)
    sheets = Sheet.query.filter_by(is_active=True).order_by(Sheet.date_created.asc()).all()
    if not sheets:
        first_sheet = Sheet(date=today_str, label=today_str)
        db.session.add(first_sheet)
        db.session.commit()
        sheets = [first_sheet]
    active_sheet_id = request.args.get('sheet_id', sheets[0].id, type=int)
    transactions = Transaction.query.filter_by(sheet_id=active_sheet_id).all()
    personnel_list = Personnel.query.order_by(Personnel.name.asc()).all()
    return render_template("home.html",
                           user=current_user,
                           transactions=transactions,
                           date_today=today_str,
                           products=inventory_items,
                           sheets=sheets,
                           active_sheet_id=active_sheet_id,
                           product_sheets=product_sheets,
                           active_product_sheet_id=active_product_sheet_id,
                           personnel_list=personnel_list)


@views.route('/add-product-sheet', methods=['POST'])
@login_required
def add_product_sheet():
    from .models import SheetProduct
    today_str = datetime.now().strftime('%m/%d/%Y')
    existing_today = ProductSheet.query.filter_by(date=today_str).first()
    if existing_today:
        return jsonify({"status": "duplicate", "message": "A table for today already exists."})
    new_sheet = ProductSheet(date=today_str, label=today_str)
    db.session.add(new_sheet)
    db.session.flush()
    prev_sheet = ProductSheet.query.filter(ProductSheet.id != new_sheet.id).order_by(ProductSheet.date_created.desc()).first()
    if prev_sheet:
        for sp in prev_sheet.sheet_products:
            new_sp = SheetProduct(sheet_id=new_sheet.id, product_id=sp.product_id, stock=sp.stock, status=sp.status, last_modified=datetime.now())
            db.session.add(new_sp)
    else:
        products = Product.query.all()
        for p in products:
            sp = SheetProduct(sheet_id=new_sheet.id, product_id=p.id, stock=0, status="Out of Stock", last_modified=datetime.now())
            db.session.add(sp)
    db.session.commit()
    return jsonify({"status": "success", "sheet_id": new_sheet.id, "label": new_sheet.label})


@views.route('/get-product-sheet/<int:sheet_id>', methods=['GET'])
@login_required
def get_product_sheet(sheet_id):
    from .models import SheetProduct
    sheet_products = SheetProduct.query.filter_by(sheet_id=sheet_id).all()
    return jsonify({
        "status": "success",
        "products": [{
            "id": sp.product_id,
            "sheet_product_id": sp.id,
            "name": sp.product.name,
            "category": sp.product.category,
            "price": sp.product.price,
            "stock": sp.stock,
            "unit": sp.product.unit,
            "status": sp.status,
            "last_modified": sp.last_modified.strftime('%Y-%m-%dT%H:%M:%S') if sp.last_modified else ''
        } for sp in sheet_products]
    })


@views.route('/get-product-stocks/<int:sheet_id>', methods=['GET'])
@login_required
def get_product_stocks(sheet_id):
    from .models import SheetProduct
    sheet_products = SheetProduct.query.filter_by(sheet_id=sheet_id).all()
    return jsonify({"status": "success", "products": [{"id": sp.product_id, "stock": sp.stock, "status": sp.status} for sp in sheet_products]})


@views.route('/export-product-sheet/<int:sheet_id>', methods=['GET'])
@login_required
def export_product_sheet(sheet_id):
    import io, openpyxl
    from flask import send_file
    from openpyxl.styles import Font, PatternFill
    sheet = ProductSheet.query.get_or_404(sheet_id)
    sheet_products = SheetProduct.query.filter_by(sheet_id=sheet_id).join(Product).order_by(Product.name.asc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet.label[:31].replace('/', '-').replace('\\', '-')
    ws.append(["#", "Product Name", "Category", "Price/Unit", "Unit", "Stock", "Status", "Last Modified"])
    for i, sp in enumerate(sheet_products, 1):
        p = sp.product
        ws.append([i, p.name, p.category, round(p.price,2), p.unit, sp.stock, sp.status, sp.last_modified.strftime('%m/%d/%Y %I:%M %p') if sp.last_modified else ''])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Products-{sheet.label.replace('/', '-')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@views.route('/add-sheet', methods=['POST'])
@login_required
def add_sheet():
    today_str = datetime.now().strftime('%m/%d/%Y')
    existing_today = Sheet.query.filter_by(date=today_str).count()
    label = today_str if existing_today == 0 else f"{today_str} ({existing_today})"
    new_sheet = Sheet(date=today_str, label=label)
    db.session.add(new_sheet)
    db.session.commit()
    return jsonify({"status": "success", "sheet_id": new_sheet.id, "label": new_sheet.label})


@views.route('/add-daily-sheet', methods=['POST'])
@login_required
def add_daily_sheet():
    data = request.get_json()
    new_date_str = data.get('date')
    existing_log = Transaction.query.filter_by(box_name=f"Log-{new_date_str}").first()
    if existing_log:
        return jsonify({"success": False, "message": "Log already exists"}), 400
    new_transaction = Transaction(box_name=f"Log-{new_date_str}", quantity=0, total_price=0.0, status="Initialized", user_id=current_user.id, sheet_id=None)
    try:
        db.session.add(new_transaction)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


@views.route('/admin')
@login_required
def admin():
    if not current_user.is_admin:
        flash("Access Denied: You do not have permission to view this page.", category='error')
        return redirect(url_for('views.home'))
    all_users = User.query.order_by(User.date_created.asc()).all()
    return render_template("admin.html", user=current_user, all_users=all_users)


# ==================== PRODUCT ADD/EDIT (FIXED) ====================

@views.route('/add-product', methods=['POST'])
@login_required
def add_product():
    from .models import SheetProduct
    from sqlalchemy.exc import IntegrityError

    name = request.form.get('name', '').strip()
    category = request.form.get('category')
    unit = request.form.get('unit')
    price_str = request.form.get('price', '').strip()
    stock_str = request.form.get('stock', '').strip()
    sheet_id = request.form.get('sheet_id', type=int)

    # Name: only letters and spaces
    if not name or not re.match(r"^[a-zA-Z\s]+$", name):
        return jsonify({"status": "error", "message": "Product name must contain only letters and spaces."}), 400
    if not price_str:
        return jsonify({"status": "error", "message": "Price is required."}), 400
    if not stock_str:
        return jsonify({"status": "error", "message": "Stock quantity is required."}), 400
    try:
        price = float(price_str)
        qty = int(stock_str)
    except ValueError:
        return jsonify({"status": "error", "message": "Invalid price or stock number."}), 400

    now = datetime.now()
    try:
        product = Product.query.filter(Product.name.ilike(name)).first()
        if product:
            product.price = price
            product.category = category
            product.unit = unit
            product.last_modified = now
        else:
            product = Product(name=name, category=category, unit=unit, price=price, date_updated=now, last_modified=now)
            db.session.add(product)
            db.session.flush()
        if sheet_id:
            existing_sp = SheetProduct.query.filter_by(sheet_id=sheet_id, product_id=product.id).first()
            if existing_sp:
                return jsonify({"status": "error", "message": f"'{name}' already exists on this sheet."}), 400
            sp = SheetProduct(sheet_id=sheet_id, product_id=product.id, stock=qty, last_modified=now)
            update_sheet_product_status(sp)
            db.session.add(sp)
        db.session.commit()
        return jsonify({"status": "success", "product_id": product.id, "last_modified": now.strftime('%Y-%m-%dT%H:%M:%S')}), 200
    except IntegrityError:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"'{name}' already exists on this sheet."}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


@views.route('/update-product/<int:product_id>/sheet/<int:sheet_id>', methods=['POST'])
@login_required
def update_sheet_product(product_id, sheet_id):
    from .models import SheetProduct
    product = Product.query.get(product_id)
    sp = SheetProduct.query.filter_by(sheet_id=sheet_id, product_id=product_id).first()
    if not product or not sp:
        return jsonify({"status": "error", "message": "Product not found."}), 404

    name = request.form.get('name', '').strip()
    if not name or not re.match(r"^[a-zA-Z\s]+$", name):
        return jsonify({"status": "error", "message": "Product name must contain only letters and spaces."}), 400

    price_str = request.form.get('price', '').strip()
    stock_str = request.form.get('stock', '').strip()
    if not price_str or not stock_str:
        return jsonify({"status": "error", "message": "Price and stock are required."}), 400
    try:
        price = float(price_str)
        stock = int(stock_str)
    except ValueError:
        return jsonify({"status": "error", "message": "Invalid price or stock number."}), 400

    category = request.form.get('category')
    unit = request.form.get('unit')
    now = datetime.now()
    try:
        product.name = name
        product.category = category
        product.unit = unit
        product.price = price
        product.date_updated = now
        product.last_modified = now
        sp.stock = stock
        sp.last_modified = now
        update_sheet_product_status(sp)
        db.session.commit()
        return jsonify({"status": "success", "product_id": product.id, "last_modified": now.strftime('%Y-%m-%dT%H:%M:%S')}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


# ==================== DELETE PRODUCT ====================

@views.route('/delete-product/<int:id>', methods=['DELETE'])
@login_required
def delete_product(id):
    from .models import SheetProduct, TransactionItem, Transaction
    product = Product.query.get(id)
    if not product:
        return jsonify({"status": "error", "message": "Product not found."}), 404
    problematic_boxes = []
    for ti in TransactionItem.query.filter_by(product_id=id):
        trans = ti.transaction
        if trans and len(trans.items) == 1:
            problematic_boxes.append(trans.box_name)
    if problematic_boxes:
        return jsonify({"status": "error", "message": f"Cannot delete '{product.name}' because it is the only product in these boxes: {', '.join(problematic_boxes)}. Remove the boxes first or add other products to them."}), 400
    try:
        SheetProduct.query.filter_by(product_id=id).delete()
        db.session.delete(product)
        db.session.commit()
        return jsonify({"status": "success"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


# ==================== TRANSACTIONS ROUTES ====================

@views.route('/transactions')
@login_required
def transactions():
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    products = Product.query.all()
    transactions = Transaction.query.filter(Transaction.date_created.contains(date_str)).all()
    return render_template("transactions.html", products=products, user=current_user, transactions=transactions, date_today=date_str)


@views.route('/get-box-details/<int:id>', methods=['GET'])
@login_required
def get_box_details(id):
    transaction = Transaction.query.get_or_404(id)
    products_list = [{'id': item.product_id, 'quantity': item.quantity, 'unit_price': item.unit_price, 'name': item.product.name if item.product else 'Unknown'} for item in transaction.items]
    return jsonify({'status': 'success', 'box_name': transaction.box_name, 'products': products_list})


@views.route('/update-box-status', methods=['POST'])
def update_box_status():
    data = request.get_json()
    box_id = data.get('id')
    new_status = data.get('status')
    transaction = Transaction.query.get(box_id)
    if transaction:
        now = datetime.now()
        transaction.status = new_status
        transaction.last_modified = now
        db.session.commit()
        return jsonify({"status": "success", "last_modified": now.strftime('%Y-%m-%dT%H:%M:%S')})
    return jsonify({"status": "error"}), 404


@views.route('/add-transaction', methods=['POST'])
@login_required
def add_transaction():
    from .models import SheetProduct
    box_name = request.form.get('box_name', '').strip()
    product_ids = request.form.getlist('product_ids[]')
    quantities = request.form.getlist('quantities[]')
    sheet_id = request.form.get('sheet_id', type=int)
    if not box_name or not product_ids or not sheet_id:
        return jsonify({"status": "error", "message": "Missing data"}), 400
    trans_sheet = Sheet.query.get(sheet_id)
    product_sheet = ProductSheet.query.filter_by(date=trans_sheet.date).first() if trans_sheet else None
    total_price = 0.0
    total_qty = 0
    items_to_add = []
    for p_id, qty_str in zip(product_ids, quantities):
        product = Product.query.get(p_id)
        if not product:
            return jsonify({"status": "error", "message": f"Product ID {p_id} not found."}), 404
        try:
            q = int(qty_str) if qty_str and str(qty_str).isdigit() else 0
        except ValueError:
            q = 0
        if q <= 0:
            continue
        if product_sheet:
            sp = SheetProduct.query.filter_by(sheet_id=product_sheet.id, product_id=product.id).first()
            available = sp.stock if sp else 0
        else:
            available = 0
        if q > available:
            return jsonify({"status": "error", "message": f"Not enough stock for '{product.name}'. Available: {available} {product.unit}, Requested: {q}."}), 400
        total_price += product.price * q
        total_qty += q
        items_to_add.append({'product': product, 'sp': sp if product_sheet else None, 'quantity': q, 'unit_price': product.price})
    if not items_to_add:
        return jsonify({"status": "error", "message": "No valid products with quantity > 0."}), 400
    try:
        now = datetime.now()
        new_trans = Transaction(box_name=box_name, quantity=total_qty, total_price=total_price, status="Initialized", user_id=current_user.id, sheet_id=sheet_id, date_created=now, last_modified=now)
        db.session.add(new_trans)
        db.session.flush()
        for item in items_to_add:
            db.session.add(TransactionItem(transaction_id=new_trans.id, product_id=item['product'].id, quantity=item['quantity'], unit_price=item['unit_price']))
            if item['sp']:
                item['sp'].stock -= item['quantity']
                item['sp'].last_modified = now
                update_sheet_product_status(item['sp'])
        db.session.commit()
        return jsonify({"status": "success", "transaction": {"id": new_trans.id, "box_name": new_trans.box_name, "quantity": new_trans.quantity, "total_price": new_trans.total_price, "status": new_trans.status, "date_created": now.strftime('%Y-%m-%dT%H:%M:%S'), "last_modified": now.strftime('%Y-%m-%dT%H:%M:%S')}})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


@views.route('/update-transaction/<int:id>', methods=['POST'])
@login_required
def update_transaction(id):
    from .models import SheetProduct
    transaction = Transaction.query.get_or_404(id)
    box_name = request.form.get('box_name', '').strip()
    product_ids = request.form.getlist('product_ids[]')
    quantities = request.form.getlist('quantities[]')
    if not box_name or not product_ids:
        return jsonify({"status": "error", "message": "Missing data"}), 400
    trans_sheet = Sheet.query.get(transaction.sheet_id)
    product_sheet = ProductSheet.query.filter_by(date=trans_sheet.date).first() if trans_sheet else None
    # restore old stock
    for old_item in transaction.items:
        if old_item.product and product_sheet:
            sp = SheetProduct.query.filter_by(sheet_id=product_sheet.id, product_id=old_item.product_id).first()
            if sp:
                sp.stock += old_item.quantity
                update_sheet_product_status(sp)
    items_to_add = []
    total_price = 0.0
    total_qty = 0
    for p_id, qty_str in zip(product_ids, quantities):
        product = Product.query.get(p_id)
        if not product:
            db.session.rollback()
            return jsonify({"status": "error", "message": f"Product ID {p_id} not found."}), 404
        try:
            q = int(qty_str) if qty_str and str(qty_str).isdigit() else 0
        except ValueError:
            q = 0
        if q <= 0:
            continue
        sp = SheetProduct.query.filter_by(sheet_id=product_sheet.id, product_id=product.id).first() if product_sheet else None
        available = sp.stock if sp else 0
        if q > available:
            db.session.rollback()
            return jsonify({"status": "error", "message": f"Not enough stock for '{product.name}'. Available: {available} {product.unit}, Requested: {q}."}), 400
        total_price += product.price * q
        total_qty += q
        items_to_add.append({'product': product, 'sp': sp, 'quantity': q, 'unit_price': product.price})
    if not items_to_add:
        db.session.rollback()
        return jsonify({"status": "error", "message": "No valid products with quantity > 0."}), 400
    try:
        TransactionItem.query.filter_by(transaction_id=id).delete()
        now = datetime.now()
        transaction.box_name = box_name
        transaction.total_price = total_price
        transaction.quantity = total_qty
        transaction.last_modified = now
        for item in items_to_add:
            db.session.add(TransactionItem(transaction_id=id, product_id=item['product'].id, quantity=item['quantity'], unit_price=item['unit_price']))
            if item['sp']:
                item['sp'].stock -= item['quantity']
                item['sp'].last_modified = now
                update_sheet_product_status(item['sp'])
        db.session.commit()
        return jsonify({"status": "success", "transaction_id": id, "last_modified": now.strftime('%Y-%m-%dT%H:%M:%S')})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


@views.route('/get-sheet-transactions/<int:sheet_id>', methods=['GET'])
@login_required
def get_sheet_transactions(sheet_id):
    transactions = Transaction.query.filter_by(sheet_id=sheet_id).all()
    return jsonify({"status": "success", "transactions": [{"id": t.id, "box_name": t.box_name, "quantity": t.quantity, "total_price": t.total_price, "status": t.status, "date_created": t.date_created.strftime('%Y-%m-%dT%H:%M:%S') if t.date_created else '', "last_modified": t.last_modified.strftime('%Y-%m-%dT%H:%M:%S') if t.last_modified else ''} for t in transactions]})


@views.route('/delete-transaction/<int:id>', methods=['DELETE'])
@login_required
def delete_transaction(id):
    from .models import SheetProduct
    transaction = Transaction.query.get(id)
    if transaction:
        try:
            trans_sheet = Sheet.query.get(transaction.sheet_id)
            product_sheet = ProductSheet.query.filter_by(date=trans_sheet.date).first() if trans_sheet else None
            for item in transaction.items:
                if item.product and product_sheet:
                    sp = SheetProduct.query.filter_by(sheet_id=product_sheet.id, product_id=item.product_id).first()
                    if sp:
                        sp.stock += item.quantity
                        sp.last_modified = datetime.now()
                        update_sheet_product_status(sp)
            db.session.delete(transaction)
            db.session.commit()
            return jsonify({"status": "success"})
        except Exception as e:
            db.session.rollback()
            return jsonify({"status": "error", "message": str(e)}), 500
    return jsonify({"status": "error", "message": "Transaction not found."}), 404


@views.route('/export-sheet/<int:sheet_id>', methods=['GET'])
@login_required
def export_sheet(sheet_id):
    import io, openpyxl
    from flask import send_file
    from openpyxl.styles import Font, PatternFill

    sheet = Sheet.query.get_or_404(sheet_id)
    transactions = Transaction.query.filter_by(sheet_id=sheet_id).all()

    wb = openpyxl.Workbook()

    # --- Sheet 1: Transactions (main list) ---
    ws1 = wb.active
    ws1.title = sheet.label[:31].replace('/', '-').replace('\\', '-')
    ws1.append(["#", "Box Name", "Total Qty", "Total Price (Php)", "Status", "Date Issued"])
    for i, t in enumerate(transactions, 1):
        ws1.append([i, t.box_name, t.quantity, round(t.total_price, 2), t.status,
                    t.date_created.strftime('%Y-%m-%d %H:%M') if t.date_created else ''])

    # --- Sheet 2: Products inside each box ---
    ws2 = wb.create_sheet("Products in Boxes")
    ws2.append(["Box Name", "Product Name", "Quantity", "Unit Price", "Total"])
    for t in transactions:
        for item in t.items:
            total = item.quantity * item.unit_price
            ws2.append([t.box_name, item.product.name, item.quantity,
                        round(item.unit_price, 2), round(total, 2)])

    # Style headers for both sheets
    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")

    # Save and send file
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{sheet.label.replace('/', '-')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ==================== PERSONNEL ROUTES ====================

@views.route('/add-personnel', methods=['POST'])
@login_required
def add_personnel():
    if not current_user.is_admin:
        return jsonify({"status": "error", "message": "Access denied."}), 403
    data = request.get_json()
    if not data:
        return jsonify({"status": "error", "message": "Invalid JSON."}), 400
    name = data.get('name', '').strip()
    phone = data.get('phone', '').strip()
    role = data.get('role', 'Delivery')
    status = data.get('status', 'Available')
    if not name or not re.match(r"^[a-zA-Z\s]+$", name):
        return jsonify({"status": "error", "message": "Name must contain only letters and spaces."}), 400
    if not phone or not re.match(r"^\d{11}$", phone):
        return jsonify({"status": "error", "message": "Phone must be exactly 11 digits."}), 400
    existing_name = Personnel.query.filter(Personnel.name.ilike(name)).first()
    if existing_name:
        return jsonify({"status": "error", "message": f"A personnel named '{name}' already exists."}), 400
    existing_phone = Personnel.query.filter(Personnel.phone == phone).first()
    if existing_phone:
        return jsonify({"status": "error", "message": "Phone number already exists."}), 400
    try:
        new_p = Personnel(name=name, phone=phone, role=role, status=status)
        db.session.add(new_p)
        db.session.commit()
        return jsonify({"status": "success", "id": new_p.id, "name": new_p.name, "phone": new_p.phone, "role": new_p.role, "status": new_p.status})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


@views.route('/update-personnel/<int:id>', methods=['POST'])
@login_required
def update_personnel(id):
    personnel = Personnel.query.get(id)
    if not personnel:
        return jsonify({"status": "error", "message": "Personnel not found."}), 404
    data = request.get_json()
    if not data:
        return jsonify({"status": "error", "message": "Invalid JSON."}), 400
    if current_user.is_admin:
        if 'name' in data:
            new_name = data.get('name', '').strip()
            if not new_name or not re.match(r"^[a-zA-Z\s]+$", new_name):
                return jsonify({"status": "error", "message": "Name must contain only letters and spaces."}), 400
            if new_name.lower() != personnel.name.lower():
                existing = Personnel.query.filter(Personnel.name.ilike(new_name)).first()
                if existing:
                    return jsonify({"status": "error", "message": f"A personnel named '{new_name}' already exists."}), 400
            personnel.name = new_name
        if 'phone' in data:
            new_phone = data.get('phone', '').strip()
            if not new_phone or not re.match(r"^\d{11}$", new_phone):
                return jsonify({"status": "error", "message": "Phone must be exactly 11 digits."}), 400
            existing_phone = Personnel.query.filter(Personnel.phone == new_phone, Personnel.id != id).first()
            if existing_phone:
                return jsonify({"status": "error", "message": "Phone number already exists."}), 400
            personnel.phone = new_phone
        if 'role' in data:
            personnel.role = data.get('role')
    if 'status' in data:
        personnel.status = data.get('status')
    try:
        db.session.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


@views.route('/delete-personnel/<int:id>', methods=['DELETE'])
@login_required
def delete_personnel(id):
    if not current_user.is_admin:
        return jsonify({"status": "error", "message": "Access denied."}), 403
    personnel = Personnel.query.get(id)
    if personnel:
        try:
            db.session.delete(personnel)
            db.session.commit()
            return jsonify({"status": "success"})
        except Exception as e:
            db.session.rollback()
            return jsonify({"status": "error", "message": str(e)}), 500
    return jsonify({"status": "error", "message": "Personnel not found."}), 404


# ==================== ADMIN PANEL ROUTES ====================

@views.route('/get-accounts', methods=['GET'])
@login_required
def get_accounts():
    if not current_user.is_admin:
        return jsonify({"status": "error", "message": "Access denied."}), 403
    users = User.query.order_by(User.date_created.asc()).all()
    return jsonify({"status": "success", "accounts": [{"id": u.id, "email": u.email, "first_name": u.first_name, "is_admin": u.is_admin, "is_online": u.is_online, "date_created": u.date_created.strftime('%Y-%m-%dT%H:%M:%S') if u.date_created else '', "last_login": u.last_login.strftime('%Y-%m-%dT%H:%M:%S') if u.last_login else '', "last_updated": u.last_updated.strftime('%Y-%m-%dT%H:%M:%S') if u.last_updated else ''} for u in users]})


@views.route('/add-account', methods=['POST'])
@login_required
def add_account():
    if not current_user.is_admin:
        return jsonify({"status": "error", "message": "Access denied."}), 403
    from werkzeug.security import generate_password_hash
    data = request.get_json()
    email = data.get('email', '').strip().lower()
    first_name = data.get('first_name', '').strip()
    password = data.get('password', '').strip()
    is_admin = bool(data.get('is_admin', False))
    if not email or not first_name or not password:
        return jsonify({"status": "error", "message": "All fields are required."}), 400
    if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
        return jsonify({"status": "error", "message": "Invalid email format."}), 400
    if len(password) < 7:
        return jsonify({"status": "error", "message": "Password must be at least 7 characters."}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({"status": "error", "message": "Email already exists."}), 400
    try:
        now = datetime.now()
        new_user = User(email=email, first_name=first_name, password=generate_password_hash(password, method='pbkdf2:sha256'), is_admin=is_admin, date_created=now, is_online=False)
        db.session.add(new_user)
        db.session.commit()
        return jsonify({"status": "success", "id": new_user.id, "email": new_user.email, "first_name": new_user.first_name, "is_admin": new_user.is_admin, "is_online": new_user.is_online, "date_created": now.strftime('%Y-%m-%dT%H:%M:%S'), "last_login": ''})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


@views.route('/update-account/<int:id>', methods=['POST'])
@login_required
def update_account(id):
    if not current_user.is_admin:
        return jsonify({"status": "error", "message": "Access denied."}), 403
    from werkzeug.security import generate_password_hash
    user = User.query.get(id)
    if not user:
        return jsonify({"status": "error", "message": "Account not found."}), 404
    data = request.get_json()
    email = data.get('email', '').strip().lower()
    first_name = data.get('first_name', '').strip()
    password = data.get('password', '').strip()
    is_admin = bool(data.get('is_admin', False))
    if not email or not first_name:
        return jsonify({"status": "error", "message": "Email and name are required."}), 400
    if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
        return jsonify({"status": "error", "message": "Invalid email format."}), 400
    duplicate = User.query.filter(User.email == email, User.id != id).first()
    if duplicate:
        return jsonify({"status": "error", "message": "Email already in use by another account."}), 400
    try:
        user.email = email
        user.first_name = first_name
        user.is_admin = is_admin
        if password:
            if len(password) < 7:
                return jsonify({"status": "error", "message": "Password must be at least 7 characters."}), 400
            user.password = generate_password_hash(password, method='pbkdf2:sha256')
        user.last_updated = datetime.now()
        db.session.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


@views.route('/delete-account/<int:id>', methods=['DELETE'])
@login_required
def delete_account(id):
    if not current_user.is_admin:
        return jsonify({"status": "error", "message": "Access denied."}), 403
    if id == current_user.id:
        return jsonify({"status": "error", "message": "You cannot delete your own account."}), 400
    user = User.query.get(id)
    if not user:
        return jsonify({"status": "error", "message": "Account not found."}), 404
    try:
        db.session.delete(user)
        db.session.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


# ==================== HISTORY & ARCHIVING ROUTES ====================

@views.route('/archive-sheet/<int:sheet_id>', methods=['POST'])
@login_required
def archive_sheet(sheet_id):
    sheet = Sheet.query.get(sheet_id)
    if sheet:
        sheet.is_active = False
        db.session.commit()
        return jsonify({"status": "success"})
    return jsonify({"status": "error"}), 404


@views.route('/archive-product-sheet/<int:sheet_id>', methods=['POST'])
@login_required
def archive_product_sheet(sheet_id):
    sheet = ProductSheet.query.get(sheet_id)
    if sheet:
        sheet.is_active = False
        db.session.commit()
        return jsonify({"status": "success"})
    return jsonify({"status": "error"}), 404


@views.route('/restore-sheet/<int:sheet_id>', methods=['POST'])
@login_required
def restore_sheet(sheet_id):
    sheet = Sheet.query.get(sheet_id)
    if sheet:
        sheet.is_active = True
        db.session.commit()
        return jsonify({"status": "success"})
    return jsonify({"status": "error"}), 404


@views.route('/restore-product-sheet/<int:sheet_id>', methods=['POST'])
@login_required
def restore_product_sheet(sheet_id):
    sheet = ProductSheet.query.get(sheet_id)
    if sheet:
        sheet.is_active = True
        db.session.commit()
        return jsonify({"status": "success"})
    return jsonify({"status": "error"}), 404


@views.route('/get-archived-sheets', methods=['GET'])
@login_required
def get_archived_sheets():
    transaction_sheets = Sheet.query.filter_by(is_active=False).order_by(Sheet.date_created.desc()).all()
    product_sheets = ProductSheet.query.filter_by(is_active=False).order_by(ProductSheet.date_created.desc()).all()
    return jsonify({"transaction_sheets": [{"id": s.id, "label": s.label, "date": s.date} for s in transaction_sheets], "product_sheets": [{"id": s.id, "label": s.label, "date": s.date} for s in product_sheets]})


@views.route('/export-history-transaction-sheet/<int:sheet_id>', methods=['GET'])
@login_required
def export_history_transaction_sheet(sheet_id):
    import io, openpyxl
    from flask import send_file
    sheet = Sheet.query.get_or_404(sheet_id)
    transactions = Transaction.query.filter_by(sheet_id=sheet_id).all()
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Transactions"
    ws1.append(["#", "Box Name", "Total Qty", "Total Price (Php)", "Status", "Date Issued", "Last Modified"])
    for i, t in enumerate(transactions, 1):
        ws1.append([i, t.box_name, t.quantity, round(t.total_price,2), t.status, t.date_created.strftime('%Y-%m-%d %H:%M') if t.date_created else '', t.last_modified.strftime('%Y-%m-%d %H:%M') if t.last_modified else ''])
    ws2 = wb.create_sheet("Products in Boxes")
    ws2.append(["Box Name", "Product Name", "Quantity", "Unit Price", "Total"])
    for t in transactions:
        for item in t.items:
            total = item.quantity * item.unit_price
            ws2.append([t.box_name, item.product.name, item.quantity, round(item.unit_price,2), round(total,2)])
    from openpyxl.styles import Font, PatternFill
    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"History_{sheet.label}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@views.route('/get-product-sheet-for-transaction-sheet/<int:sheet_id>', methods=['GET'])
@login_required
def get_product_sheet_for_transaction_sheet(sheet_id):
    sheet = Sheet.query.get(sheet_id)
    if not sheet:
        return jsonify({"error": "Sheet not found"}), 404
    product_sheet = ProductSheet.query.filter_by(date=sheet.date).first()
    if product_sheet:
        return jsonify({"product_sheet_id": product_sheet.id})
    return jsonify({"product_sheet_id": None})


@views.route('/get-sheet-date/<int:sheet_id>', methods=['GET'])
@login_required
def get_sheet_date(sheet_id):
    sheet = Sheet.query.get(sheet_id)
    if sheet:
        return jsonify({"date": sheet.date})
    return jsonify({"date": None}), 404


@views.route('/get-product-sheet-id-by-date', methods=['GET'])
@login_required
def get_product_sheet_id_by_date():
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({"product_sheet_id": None}), 400
    product_sheet = ProductSheet.query.filter_by(date=date_str, is_active=True).first()
    if not product_sheet:
        product_sheet = ProductSheet.query.filter_by(date=date_str).first()
    return jsonify({"product_sheet_id": product_sheet.id if product_sheet else None})


# ==================== PERMANENT DELETE FOR ARCHIVED SHEETS ====================

@views.route('/delete-archived-sheet/<int:sheet_id>', methods=['DELETE'])
@login_required
def delete_archived_sheet(sheet_id):
    if not current_user.is_admin:
        return jsonify({"status": "error", "message": "Access denied."}), 403
    sheet = Sheet.query.get(sheet_id)
    if not sheet:
        return jsonify({"status": "error", "message": "Sheet not found."}), 404
    try:
        db.session.delete(sheet)          # Cascade will delete transactions and items
        db.session.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    
@views.route('/delete-archived-product-sheet/<int:sheet_id>', methods=['DELETE'])
@login_required
def delete_archived_product_sheet(sheet_id):
    if not current_user.is_admin:
        return jsonify({"status": "error", "message": "Access denied."}), 403
    sheet = ProductSheet.query.get(sheet_id)
    if not sheet:
        return jsonify({"status": "error", "message": "Product sheet not found."}), 404
    try:
        # Delete related sheet_products and snapshots (cascade should handle)
        db.session.delete(sheet)
        db.session.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500